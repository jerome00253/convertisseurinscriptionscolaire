
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkinter import ttk
from PIL import Image, ImageTk
from tkcalendar import DateEntry
import os
import shutil
from datetime import datetime
import threading
import sys
import re

# For Excel styling and charts
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class ExcelConverterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Convertisseur inscription scolaire")
        self.root.geometry("1100x850")
        
        # Color palette
        self.bg_color = "#f8f9fa"
        self.primary_color = "#004a99" # Illzach Blue
        self.secondary_color = "#ffcc00" # Illzach Yellow
        self.text_color = "#202124"
        
        self.root.configure(bg=self.bg_color)
        
        # Internal state
        self.source_file = None
        self.available_sheets = []
        self.sheet_vars = {} 
        self.school_vars = {} 
        self.date_filter_active = tk.BooleanVar(value=False)
        # Summary Generation Toggle
        self.gen_summary_var = tk.BooleanVar(value=False)

        # UI Elements
        self.setup_ui()

    def setup_ui(self):
        # Header Frame
        header_frame = tk.Frame(self.root, bg="white", height=100)
        header_frame.pack(fill="x")
        
        try:
            logo_path = resource_path("1280px-LogoIllzach.jpg")
            img = Image.open(logo_path)
            aspect_ratio = img.width / img.height
            new_height = 70
            new_width = int(new_height * aspect_ratio)
            img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.logo_img = ImageTk.PhotoImage(img)
            logo_label = tk.Label(header_frame, image=self.logo_img, bg="white")
            logo_label.pack(side="left", padx=20, pady=10)
        except Exception as e:
            print(f"Could not load logo: {e}")

        title_label = tk.Label(header_frame, text="Convertisseur Inscription Scolaire", 
                               bg="white", fg=self.primary_color, 
                               font=("Segoe UI", 22, "bold"))
        title_label.pack(side="left", padx=10)

        # Main frame
        main_frame = tk.Frame(self.root, bg=self.bg_color, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)

        # Left Column: File & Filters
        left_col = tk.Frame(main_frame, bg=self.bg_color)
        left_col.pack(side="left", fill="both", expand=True, padx=(0, 15))

        # 1. File Selection
        file_frame = tk.LabelFrame(left_col, text="1. Sélection du fichier", bg=self.bg_color, font=("Segoe UI", 10, "bold"))
        file_frame.pack(fill="x", pady=(0, 15))
        
        self.select_btn = tk.Button(file_frame, text="📁 Choisir l'export Illzach", 
                                   command=self.select_file, bg=self.primary_color, fg="white",
                                   font=("Segoe UI", 10), relief="flat", padx=15, pady=8, cursor="hand2")
        self.select_btn.pack(pady=10)
        
        self.file_label = tk.Label(file_frame, text="Aucun fichier sélectionné", 
                                  bg=self.bg_color, fg=self.text_color, font=("Segoe UI", 9, "italic"), wraplength=400)
        self.file_label.pack(pady=(0, 10))

        # 2. Advanced Filters
        filter_frame = tk.LabelFrame(left_col, text="2. Filtres avancés", bg=self.bg_color, font=("Segoe UI", 10, "bold"))
        filter_frame.pack(fill="x", pady=0)

        # Date Filter Toggle
        toggle_frame = tk.Frame(filter_frame, bg=self.bg_color)
        toggle_frame.pack(fill="x", padx=10, pady=(10, 0))
        tk.Checkbutton(toggle_frame, text="Activer le filtrage par date", variable=self.date_filter_active, 
                       command=self.update_ui_states, bg=self.bg_color, font=("Segoe UI", 9, "bold")).pack(side="left")

        # Date Filter Sub-frame
        self.date_subframe = tk.Frame(filter_frame, bg=self.bg_color)
        self.date_subframe.pack(fill="x", padx=10, pady=10)
        
        tk.Label(self.date_subframe, text="Date de début :", bg=self.bg_color).grid(row=0, column=0, sticky="w", pady=5)
        self.start_date_entry = DateEntry(self.date_subframe, width=12, background='darkblue',
                                         foreground='white', borderwidth=2, locale='fr_FR', date_pattern='dd/mm/yyyy')
        self.start_date_entry.grid(row=0, column=1, padx=10, sticky="w")
        
        tk.Label(self.date_subframe, text="Date de fin :", bg=self.bg_color).grid(row=1, column=0, sticky="w", pady=5)
        self.end_date_entry = DateEntry(self.date_subframe, width=12, background='darkblue',
                                       foreground='white', borderwidth=2, locale='fr_FR', date_pattern='dd/mm/yyyy')
        self.end_date_entry.grid(row=1, column=1, padx=10, sticky="w")

        # Derogation Filter Sub-frame
        derog_subframe = tk.Frame(filter_frame, bg=self.bg_color)
        derog_subframe.pack(fill="x", padx=10, pady=(0, 15))
        
        tk.Label(derog_subframe, text="Besoin d'une dérogation ?", bg=self.bg_color).pack(side="left")
        self.derog_filter_var = tk.StringVar(value="Tous")
        self.derog_combo = ttk.Combobox(derog_subframe, textvariable=self.derog_filter_var, 
                                       values=["Tous", "Oui", "Non"], state="readonly", width=10)
        self.derog_combo.pack(side="left", padx=10)

        # Summary Generation Toggle
        summary_toggle_frame = tk.Frame(filter_frame, bg=self.bg_color)
        summary_toggle_frame.pack(fill="x", padx=10, pady=(0, 5))
        tk.Checkbutton(summary_toggle_frame, text="Générer la page de synthèse", variable=self.gen_summary_var, 
                       command=self.update_ui_states, bg=self.bg_color, font=("Segoe UI", 9, "bold")).pack(side="left")

        self.update_ui_states()

        # Middle Column: Sheet Selection
        mid_col = tk.Frame(main_frame, bg=self.bg_color)
        mid_col.pack(side="left", fill="both", expand=True, padx=(15, 15))

        sheet_frame = tk.LabelFrame(mid_col, text="3. Sélection des onglets", bg=self.bg_color, font=("Segoe UI", 10, "bold"))
        sheet_frame.pack(fill="both", expand=True)

        self.sheets_canvas = tk.Canvas(sheet_frame, bg="white", borderwidth=0, highlightthickness=0)
        self.sheets_scrollbar = ttk.Scrollbar(sheet_frame, orient="vertical", command=self.sheets_canvas.yview)
        self.sheets_list_frame = tk.Frame(self.sheets_canvas, bg="white")

        self.sheets_list_frame.bind("<Configure>", lambda e: self.sheets_canvas.configure(scrollregion=self.sheets_canvas.bbox("all")))
        self.sheets_canvas.create_window((0, 0), window=self.sheets_list_frame, anchor="nw")
        self.sheets_canvas.configure(yscrollcommand=self.sheets_scrollbar.set)

        self.sheets_canvas.pack(side="left", fill="both", expand=True)
        self.sheets_scrollbar.pack(side="right", fill="y")

        # Right Column: School Selection
        right_col = tk.Frame(main_frame, bg=self.bg_color)
        right_col.pack(side="right", fill="both", expand=True, padx=(15, 0))

        school_list_frame_container = tk.LabelFrame(right_col, text="4. Sélection des écoles", bg=self.bg_color, font=("Segoe UI", 10, "bold"))
        school_list_frame_container.pack(fill="both", expand=True)

        self.schools_canvas = tk.Canvas(school_list_frame_container, bg="white", borderwidth=0, highlightthickness=0)
        self.schools_scrollbar = ttk.Scrollbar(school_list_frame_container, orient="vertical", command=self.schools_canvas.yview)
        self.schools_list_frame = tk.Frame(self.schools_canvas, bg="white")

        self.schools_list_frame.bind("<Configure>", lambda e: self.schools_canvas.configure(scrollregion=self.schools_canvas.bbox("all")))
        self.schools_canvas.create_window((0, 0), window=self.schools_list_frame, anchor="nw")
        self.schools_canvas.configure(yscrollcommand=self.schools_scrollbar.set)

        self.schools_canvas.pack(side="left", fill="both", expand=True)
        self.schools_scrollbar.pack(side="right", fill="y")

        # Bottom section: Logs & Action
        bottom_frame = tk.Frame(self.root, bg=self.bg_color, padx=20)
        bottom_frame.pack(fill="x", pady=(10, 20))

        self.log_area = scrolledtext.ScrolledText(bottom_frame, height=6, font=("Consolas", 9), bg="white")
        self.log_area.pack(fill="x", pady=(0, 15))

        self.convert_btn = tk.Button(bottom_frame, text="🚀 GÉNÉRER L'EXPORT VÉROUILLÉ ET MIS EN FORME", 
                                    command=self.start_conversion, bg="#28a745", fg="white",
                                    font=("Segoe UI", 12, "bold"), relief="flat", padx=40, pady=15,
                                    state="disabled", cursor="hand2")
        self.convert_btn.pack()

        self.log("Application v9 prête (Synthèse Avancée).")

    def update_ui_states(self):
        # Date Filter
        date_state = "normal" if self.date_filter_active.get() else "disabled"
        for child in self.date_subframe.winfo_children():
            try:
                child.configure(state=date_state)
            except:
                pass
        
        pass
    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_area.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_area.see(tk.END)
        self.root.update_idletasks()

    def select_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.source_file = file_path
            self.file_label.config(text=os.path.basename(file_path), font=("Segoe UI", 9, "normal"))
            self.log(f"Fichier chargé : {os.path.basename(file_path)}")
            self.load_metadata()

    def load_metadata(self):
        try:
            for widget in self.sheets_list_frame.winfo_children(): widget.destroy()
            for widget in self.schools_list_frame.winfo_children(): widget.destroy()
            self.sheet_vars = {}
            self.school_vars = {}

            self.log("Analyse du fichier en cours...")
            wb = load_workbook(self.source_file, read_only=True)
            self.available_sheets = wb.sheetnames
            wb.close()

            for sheet in self.available_sheets:
                var = tk.BooleanVar(value=True)
                self.sheet_vars[sheet] = var
                cb = tk.Checkbutton(self.sheets_list_frame, text=sheet, variable=var, 
                                   bg="white", font=("Segoe UI", 9), anchor="w")
                cb.pack(fill="x", padx=5, pady=2)

            threading.Thread(target=self.extract_schools, daemon=True).start()
            
        except Exception as e:
            self.log(f"Erreur Lecture : {e}")

    def extract_schools(self):
        try:
            schools = set()
            for sheet in self.available_sheets:
                df = pd.read_excel(self.source_file, sheet_name=sheet, usecols=["Ecole"])
                if not df.empty and "Ecole" in df.columns:
                    schools.update(df["Ecole"].dropna().unique())
            
            sorted_schools = sorted([str(s) for s in schools])
            for school in sorted_schools:
                var = tk.BooleanVar(value=True)
                self.school_vars[school] = var
                cb = tk.Checkbutton(self.schools_list_frame, text=school, variable=var, 
                                   bg="white", font=("Segoe UI", 9), anchor="w")
                cb.pack(fill="x", padx=5, pady=2)
            
            self.convert_btn.config(state="normal")
            self.log(f"Analyse terminée : {len(self.available_sheets)} onglets, {len(sorted_schools)} écoles.")
        except Exception as e:
            self.log(f"Erreur Analyse Écoles : {e}")

    def start_conversion(self):
        start_date = None
        end_date = None
        
        if self.date_filter_active.get():
            try:
                sd_str = self.start_date_entry.get()
                if sd_str:
                    start_date = datetime.strptime(sd_str, "%d/%m/%Y")
                ed_str = self.end_date_entry.get()
                if ed_str:
                    end_date = datetime.strptime(ed_str, "%d/%m/%Y")
            except Exception:
                messagebox.showerror("Erreur Date", "Format de date invalide.")
                return

        selected_sheets = [s for s, var in self.sheet_vars.items() if var.get()]
        selected_schools = [s for s, var in self.school_vars.items() if var.get()]
        
        if not selected_sheets:
            messagebox.showwarning("Onglets", "Sélectionnez au moins un onglet.")
            return
        if not selected_schools:
            messagebox.showwarning("Écoles", "Sélectionnez au moins une école.")
            return

        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Export_Illzach_EMS_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
        if not output_path: return

        self.convert_btn.config(state="disabled")
        
        params = {
            "output_path": output_path,
            "selected_sheets": selected_sheets,
            "selected_schools": selected_schools,
            "start_date": start_date,
            "end_date": end_date,
            "derog_filter": self.derog_filter_var.get(),
            "apply_date_filter": self.date_filter_active.get(),
            "gen_summary": self.gen_summary_var.get()
        }
        
        threading.Thread(target=self.process_conversion, args=(params,), daemon=True).start()

    def process_conversion(self, params):
        try:
            self.log("Démarrage v9 (Synthèse Avancée)...")
            temp_path = "temp_v9.xlsx"
            shutil.copy2(self.source_file, temp_path)
            
            combined_data = []
            for sheet in params["selected_sheets"]:
                df = pd.read_excel(temp_path, sheet_name=sheet)
                if df.empty: continue
                df['Onglet'] = sheet
                combined_data.append(df)
            
            if not combined_data: raise Exception("Aucune donnée.")
            df_full = pd.concat(combined_data, ignore_index=True)
            
            # --- FILTERS ---
            if params["apply_date_filter"] and 'Date de création' in df_full.columns:
                df_full['Date_dt'] = pd.to_datetime(df_full['Date de création'], format='%d/%m/%Y %H:%M:%S', errors='coerce')
                if params["start_date"]:
                    df_full = df_full[df_full['Date_dt'] >= params["start_date"]]
                if params["end_date"]:
                    limit = params["end_date"].replace(hour=23, minute=59, second=59)
                    df_full = df_full[df_full['Date_dt'] <= limit]
                df_full = df_full.drop(columns=['Date_dt'])

            if params["derog_filter"] != "Tous" and "Besoin d'une dérogation" in df_full.columns:
                df_full = df_full[df_full["Besoin d'une dérogation"].astype(str).str.lower() == params["derog_filter"].lower()]

            if "Ecole" in df_full.columns:
                df_full = df_full[df_full["Ecole"].astype(str).isin(params["selected_schools"])]

            self.log(f"Lignes retenues : {len(df_full)}")

            # Identify Status Column
            status_col = None
            for c in ["État", "Etat", "Statut", "Etat dossier"]:
                if c in df_full.columns:
                    status_col = c
                    break
            
            # --- PREPARE DATA FOR INDIVIDUAL TABS ---
            mapping = {
                'Onglet': 'Onglet', 'N° de dossier': 'N°', 'Nom enfant': 'Nom enfant', 'Prénom enfant': 'Prénom enfant',
                'Date de naissance enfant': 'Date de naissance enfant', "Besoin d'une dérogation": "Besoin d'une dérogation",
                'Adresse indiquée': 'Adresse indiquée', 'Ecole': 'Ecole', 'Classe': 'Classe', 'Cursus': 'Cursus',
                'Resp. 1 civilité': 'Resp. 1 civilité', 'Resp. 1 nom de naissance': 'Resp. 1 nom de naissance',
                "Resp. 1 nom d'usage": "Resp. 1 nom d'usage", 'Resp. 1 prénom': 'Resp. 1 prénom',
                'Resp. 1 téléphone': 'Resp. 1 téléphone', 'Resp. 1 email': 'Resp. 1 email', 'Resp. 1 adresse': 'Resp. 1 adresse',
                'Fratrie 1 nom': 'Fratrie 1 nom', 'Fratrie 1 prénom': 'Fratrie 1 prénom', 'Fratrie 1 école': 'Fratrie 1 école',
                'Fratrie 1 classe': 'Fratrie 1 classe', 'Dérogation école voulue': 'Dérogation école voulue',
                'Dérogation autre école voulue - nom': 'Dérogation autre école voulue - nom', 'Dérogation raison': 'Dérogation raison'
            }
            if status_col: mapping[status_col] = "État"
            
            available_cols = [col for col in mapping.keys() if col in df_full.columns]
            df_mapped = df_full[available_cols].rename(columns=mapping)
            if 'Ecole' in df_mapped.columns: df_mapped = df_mapped.sort_values(by='Ecole')
            
            # --- START WRITING ---
            with pd.ExcelWriter(params["output_path"], engine='openpyxl') as writer:
                # 1. ADVANCED SYNTHESIS DATA
                if params.get("gen_summary", False) and 'Ecole' in df_mapped.columns:
                    self.log("Génération de la synthèse avancée...")
                    stats = []
                    for school in df_mapped['Ecole'].unique():
                        df_s = df_mapped[df_mapped['Ecole'] == school].copy()
                        total = len(df_s)
                        
                        school_stats = {
                            'Ecole': school,
                            'Total': total
                        }
                        
                        for sheet_name in params["selected_sheets"]:
                            count = 0
                            if "Onglet" in df_s.columns:
                                count = len(df_s[df_s["Onglet"] == sheet_name])
                            school_stats[sheet_name] = count
                        
                        derog = 0
                        if "Besoin d'une dérogation" in df_s.columns:
                            derog = len(df_s[df_s["Besoin d'une dérogation"].astype(str).str.lower() == "oui"])
                        
                        perc = (derog / total * 100) if total > 0 else 0
                        school_stats['Dérogations'] = derog
                        school_stats['% Dérog.'] = round(perc, 1)
                        
                        stats.append(school_stats)
                    
                    df_stats = pd.DataFrame(stats)
                    df_stats.to_excel(writer, sheet_name="Synthèse", index=False, startrow=2)
                
                # 2. SCHOOL TABS
                if 'Ecole' in df_mapped.columns:
                    for school in df_mapped['Ecole'].unique():
                        name = str(school)[:31].replace('/', '-').replace('\\', '-')
                        if not name or name.lower() == 'nan': name = "Sans Ecole"
                        df_s = df_mapped[df_mapped['Ecole'] == school].copy()
                        df_s['N°'] = range(1, len(df_s) + 1)
                        if 'Onglet' in df_s.columns:
                            cols = df_s.columns.tolist()
                            cols.insert(0, cols.pop(cols.index('Onglet')))
                            df_s = df_s[cols]
                        df_s.to_excel(writer, sheet_name=name, index=False)
                else:
                    df_mapped['N°'] = range(1, len(df_mapped) + 1)
                    df_mapped.to_excel(writer, sheet_name="Export", index=False)
            
            # --- POST-PROCESSING STYLING & CHARTS ---
            self.log("Mise en forme et graphiques empilés...")
            wb = load_workbook(params["output_path"])
            
            # Global styles
            h_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            h_font = Font(bold=True)
            top_align = Alignment(vertical="top")
            top_align_wrap = Alignment(vertical="top", wrap_text=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Stylize Synthesis
            if "Synthèse" in wb.sheetnames:
                ws = wb["Synthèse"]
                ws.insert_rows(1, 1)
                ws["A1"] = "TABLEAU DE BORD DES INSCRIPTIONS SCOLAIRES"
                ws["A1"].font = Font(size=18, bold=True, color="004a99")
                
                last_row = ws.max_row
                last_col = ws.max_column
                
                # Style header
                for cell in ws[3]:
                    cell.fill = h_fill
                    cell.font = h_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
                
                # Style data rows
                for row in range(4, last_row + 1):
                    for col in range(1, last_col + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center")
                
                # Index of columns for Chart
                col_indices = {ws.cell(row=3, column=c).value: c for c in range(1, last_col + 1)}

                # Adjust widths
                ws.column_dimensions['A'].width = 35
                for col_idx in range(2, last_col + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 12

                # Add Total row
                t_row = last_row + 1
                ws.cell(row=t_row, column=1, value="TOTAL GÉNÉRAL").font = Font(bold=True)
                # Calculate sums for all columns except Ecole and % Dérog.
                for c_idx in range(2, last_col + 1):
                    col_str = str(ws.cell(row=3, column=c_idx).value)
                    if "%" in col_str and "rog" in col_str:
                        if "Total" in col_indices and "Dérogations" in col_indices:
                            t_col = get_column_letter(col_indices["Total"])
                            d_col = get_column_letter(col_indices["Dérogations"])
                            ws.cell(row=t_row, column=c_idx, value=f"=IF({t_col}{t_row}>0, {d_col}{t_row}/{t_col}{t_row}*100, 0)").font = Font(bold=True)
                        else:
                            c_letter = get_column_letter(c_idx)
                            ws.cell(row=t_row, column=c_idx, value=f"=AVERAGE({c_letter}4:{c_letter}{last_row})").font = Font(bold=True)
                        ws.cell(row=t_row, column=c_idx).number_format = '0.0'
                        continue
                        
                    c_letter = get_column_letter(c_idx)
                    ws.cell(row=t_row, column=c_idx, value=f"=SUM({c_letter}4:{c_letter}{last_row})").font = Font(bold=True)
                
                for col in range(1, last_col + 1):
                     ws.cell(row=t_row, column=col).border = border

                # Add STACKED Bar Chart
                # Data for chart: dynamic sheets
                stacked_cols = params["selected_sheets"]
                chart_col_indices = [col_indices[c] for c in stacked_cols if c in col_indices]
                
                if chart_col_indices:
                    chart = BarChart()
                    chart.type = "col"
                    chart.grouping = "stacked"
                    chart.overlap = 100
                    chart.title = "Répartition des États par École"
                    chart.y_axis.title = "Nombre de dossiers"
                    chart.x_axis.title = "Écoles"
                    
                    # We need to create multiple series if columns are not contiguous, 
                    # but usually they are. For simplicity, we'll try to find the range.
                    min_c = min(chart_col_indices)
                    max_c = max(chart_col_indices)
                    
                    data = Reference(ws, min_col=min_c, min_row=3, max_row=last_row, max_col=max_c)
                    cats = Reference(ws, min_col=1, min_row=4, max_row=last_row)
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    
                    chart.legend.position = "r"
                    ws.add_chart(chart, f"A{t_row + 3}")

            # Stylize School Tabs
            for sheet_name in wb.sheetnames:
                if sheet_name == "Synthèse": continue
                ws = wb[sheet_name]
                ws.freeze_panes = 'A2'
                ws.auto_filter.ref = ws.dimensions
                for row in ws.iter_rows():
                    for cell in row: cell.alignment = top_align
                for cell in ws[1]:
                    cell.fill = h_fill
                    cell.font = h_font
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                
                for col_ws in ws.columns:
                    letter = col_ws[0].column_letter
                    h = str(col_ws[0].value).lower()
                    if "dérogation raison" in h:
                        ws.column_dimensions[letter].width = 75
                        for cell in col_ws: cell.alignment = top_align_wrap
                    elif "adresse" in h:
                        ws.column_dimensions[letter].width = 45
                        for cell in col_ws: cell.alignment = top_align_wrap
                    elif "nom" in h or "prénom" in h or "ecole" in h:
                        ws.column_dimensions[letter].width = 25
                    elif h == "n°":
                        ws.column_dimensions[letter].width = 6
                        for cell in col_ws: cell.alignment = Alignment(horizontal="center", vertical="top")
                    else:
                        ws.column_dimensions[letter].width = 18

            wb.save(params["output_path"])
            if os.path.exists(temp_path): os.remove(temp_path)
            self.log("Terminé avec succès !")
            messagebox.showinfo("Succès", f"Fichier v9 généré (Synthèse avancée + Graphique empilé).")
            
        except Exception as e:
            self.log(f"ERREUR : {e}")
            messagebox.showerror("Erreur", str(e))
        finally:
            self.convert_btn.config(state="normal")
            self.select_btn.config(state="normal")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelConverterApp(root)
    root.mainloop()
